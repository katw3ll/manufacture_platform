import qrcode
from django.db import models
from colorfield.fields import ColorField
from djmoney.models.fields import MoneyField
import requests
import json
import xml.etree.ElementTree as ET
from openpyxl.reader.excel import load_workbook
from qrcode import make, QRCode
from qrcode.exceptions import DataOverflowError
import base64
import zlib


from django.contrib.auth.models import User


import sys
import uuid


def str_to_class(stro):
    try:
        return getattr(sys.modules[__name__], stro)
    except:
        return False


# Создайте свои модели здесь.

class userProfile(models.Model):
    user = models.OneToOneField(User, on_delete=models.CASCADE, related_name="profile")
    description = models.TextField(blank=True, null=True)
    location = models.CharField(max_length=30, blank=True)
    date_joined = models.DateTimeField(auto_now_add=True)
    updated_on = models.DateTimeField(auto_now=True)
    is_organizer = models.BooleanField(default=False)

    def __str__(self):
        return self.user.username



class Shelf(models.Model):

    box_number = models.BigIntegerField("box_number",
                                  help_text="",
                                  default="")

    cell_number = models.BigIntegerField("cell_number",
                                  help_text="")


    def __str__(self):
        return f"{self.box_number} || {self.cell_number}"

    class Meta:
        verbose_name = "Полка на склад"
        verbose_name_plural = "Полки на складе"


class Status(models.Model):
    name_status = models.CharField("name_status", max_length=200,
                                   help_text="Введите название статуса проекта",
                                   default="")


    def __str__(self):
        return self.name_status

    class Meta:
        verbose_name = "Статус"
        verbose_name_plural = "Статусы"



class DIM(models.Model):

    dim = models.CharField("DIM", max_length=225,
                                help_text="Введите единицу измерения, который будет отображаться у пользователя")
    can_cut = models.BooleanField("can_cut", default=False)

    def __str__(self):
        return self.dim

    class Meta:
        verbose_name = "Единица измерения"
        verbose_name_plural = "Единица измерения"


class Classes(models.Model):

    class_name = models.CharField("class_name", max_length=225,
                                  help_text="Введите название класса материалов, который будет отображаться у пользователя")

    dim = models.ForeignKey(DIM, verbose_name="Единица измерения", on_delete=models.CASCADE)

    def __str__(self):
        return self.class_name

    class Meta:
        verbose_name = "Класс материалов"
        verbose_name_plural = "Класс материалов"


class Materials(models.Model):

    material_name = models.CharField("material_name", max_length=40,
                                     help_text="Введите название материала, который будет отображаться у пользователя")
    class_id = models.ForeignKey(Classes, verbose_name="Класс материала", on_delete=models.CASCADE)
    weight = models.FloatField("weight", help_text="Введите вес условной единицы материала")

    default_lenght = models.IntegerField("default_lenght", help_text="Введите стандартную длину материала в мм",
                                         default=0)
    good_cut = models.IntegerField("good_cut", help_text="Введите длину ликвидного остатка в мм",
                                   default=0)


    def __str__(self):
        return self.material_name

    class Meta:
        verbose_name = "Название материала"
        verbose_name_plural = "Название материала"


class Colors(models.Model):

    color_name = models.CharField("color_name", max_length=225,
                                  help_text="Введите название цвета, который будет отображаться у пользователя",
                                  default='')

    color_short = models.CharField("color_short", max_length=225,
                                   help_text="Введите краткое название цвета, который будет отображаться у пользователя",
                                   default='')
    
    rgb = ColorField(default='#000000')
    ral = models.BigIntegerField("RAL", help_text="Цвет по RAL", default=0)

    color_alutech = models.IntegerField("color_alutech", help_text="Введите код цвета", default=0)


    def __str__(self):
        return self.color_name

    class Meta:
        verbose_name = "Название цвета"
        verbose_name_plural = "Название цвета"


class Parts(models.Model):

    material_id = models.ForeignKey(Materials, verbose_name="Материал", on_delete=models.CASCADE)

    artnumber = models.BigIntegerField("artnumber",
                                       help_text="Артикул", default=0)

    color_id = models.ForeignKey(Colors, verbose_name="Цвет", on_delete=models.CASCADE, null=True)
    price = MoneyField(max_digits=14, decimal_places=2, default_currency='RUB')

    #composition = models.ForeignKey(Composition, verbose_name="composition", on_delete=models.CASCADE)

    def __str__(self):
        return f"{self.material_id} || {self.color_id}"

    class Meta:
        verbose_name = "Артикул"
        verbose_name_plural = "Артикул"




class Stock(models.Model):


    quantity = models.BigIntegerField("quantity",
                                      help_text="")

    length = models.BigIntegerField("length",
                                    help_text="")

    barcode = models.BigIntegerField("barcode",
                                     help_text="")

    shelf = models.ForeignKey(Shelf, verbose_name="Shelf", on_delete=models.CASCADE)

    parts = models.ForeignKey(Parts, verbose_name="parts", on_delete=models.CASCADE)


    image = models.ImageField(upload_to="%Y/%m/%d", blank=False, default="0", verbose_name="Qr-code")


    class Meta:
        verbose_name = "Склад"
        verbose_name_plural = "Склады"

    def save(self, *args, **kwargs):

        parts = Parts.objects.filter(id=int(self.parts_id))
        from .serializers import PartsSerializer
        serializer = PartsSerializer(parts, many=True)
        # return Response(serializer.data)
        shelf = Shelf.objects.filter(id=self.shelf_id)

        str_data = str(serializer.data)

        long_text_compressed = zlib.compress(str_data.encode('utf-8'))

        report_encoded = base64.b64encode(long_text_compressed)

        report_encoded = str(report_encoded).replace("b'", "")
        report_encoded = report_encoded.replace("'", "")


        qr_code = {
            "length": self.length,
            "box_number": shelf[0].box_number,
            "cell_number": shelf[0].cell_number,
            "parts": report_encoded
        }  # Длина, цвет, название материала, артикуль, номер полки
        js = json.dumps(qr_code)
        img = qrcode.make(js)
        #filename = f'media/qr-code/g.png'
        filename = 'media/qr-code/' + str(uuid.uuid4()) + '.png'
        img.save(filename)
        #self.qr_code_id = qr_id
        self.image = filename.replace("media/", "")
        super(Stock, self).save(*args, **kwargs)



class Project(models.Model):

    receipt_date = models.DateField(verbose_name="receipt_date", help_text="Дата поступления")

    deadline_date = models.DateField(verbose_name="deadline_date", help_text="Дата дедлайна")


    #composition = models.ManyToManyField(Composition, verbose_name="Composition")

    status = models.ForeignKey(Status, verbose_name="Status", on_delete=models.CASCADE)


    def __str__(self):
        return f"№ {self.id} || {self.receipt_date}"

    class Meta:
        verbose_name = "Проект"
        verbose_name_plural = "Проекты"


class Rollets(models.Model):
    # type_roller = models.BigIntegerField("type_roller",
    #                                      help_text="Тип рольставни", default=0)

    width = models.BigIntegerField("width",
                                   help_text="Ширина", default=0)

    height = models.BigIntegerField("height",
                                    help_text="Высота", default=0)

    project = models.ForeignKey(Project, verbose_name="project", on_delete=models.CASCADE, null=True)

    parts = models.ForeignKey(Parts, verbose_name="parts", on_delete=models.CASCADE, null=True)

    status_packed = models.BooleanField("status_packed", default=False)


    def __str__(self):
        return f"{self.width} || {self.height}"

    class Meta:
        verbose_name = "Роллета"
        verbose_name_plural = "Роллеты"


class Composition(models.Model):
    # partno = models.BigIntegerField("partno",
    #                                      help_text="")

    length = models.BigIntegerField("length",
                                    help_text="")

    need_count = models.BigIntegerField("need_count",
                                        help_text="")

    quantity = models.BigIntegerField("quantity",
                                      help_text="")

    parts = models.ForeignKey(Parts, verbose_name="parts", on_delete=models.CASCADE)

    rollets = models.ForeignKey(Rollets, verbose_name="rollets", on_delete=models.CASCADE, null=True)


    def __str__(self):
        return f"{self.length} || {self.need_count}"

    class Meta:
        verbose_name = "Распил"
        verbose_name_plural = "Распилы"


class Queue(models.Model):
    quantity = models.BigIntegerField("quantity",
                                      help_text="Количество")

    length = models.BigIntegerField("length",
                                    help_text="Длинна")

    parts = models.ForeignKey(Parts, verbose_name="parts", on_delete=models.CASCADE)

    project = models.ManyToManyField(Project, verbose_name="project")


    def __str__(self):
        return f"{self.parts} || {self.project}"

    class Meta:
        verbose_name = "Связь очереди с данными"
        verbose_name_plural = "Связь очереди с данными"


class CollectionQueue(models.Model):


    queue = models.ManyToManyField(Queue, verbose_name="Queue")


    def __str__(self):
        return f"{self.queue}"

    class Meta:
        verbose_name = "Очередь"
        verbose_name_plural = "Очередь"


class Documents(models.Model):
    name = models.CharField("Имя файла", max_length=200,
                            help_text="Введите адекватное название файла", default="")

    file = models.FileField(upload_to="%Y/%m/%d", default="0", verbose_name="Файл",
                            help_text="Загрузите файл в формате txt")

    name_dow = models.CharField("", max_length=200,
                                help_text="Введите название класса", default="")


    def __str__(self):
        return self.name

    class Meta:
        verbose_name = "Загрузка документа"
        verbose_name_plural = "Загрузки документов"

    def save(self, *args, **kwargs):
        super(Documents, self).save(*args, **kwargs)
        #cl = str_to_class(self.name_dow)
        sl = {
            "colors": "add_colors",
            "dim": "add_dim",
            "parts": "add_parts",
            "classes": "add_classes",
            "materials": "aadd_materials",
            "composition": "add_composition",
            "rollets": "add_rolls",
            "stock": "add_stock"
        }
        sp = []
        if "xml" in str(self.file):
            tree = ET.parse(f'media/{self.file}')
            root = tree.getroot()
            rolletes = root[0][0][0]

            rolls = []

            for r in rolletes:
                #print(r.find('Roll_Parts'))
                #print(int(float(r.get("WIDTH").replace(",", ".")) * 1000),
                      #int(float(r.get("HEIGHT").replace(",", ".")) * 1000))
                cutting_data = []
                max_cnt = 0
                material_partno = 0
                for part in r.find('Roll_Parts'):
                    partno = int(part.get("CODE_"))
                    count = int(part.get("CNT"))
                    l = int(float(part.get("SZ").replace(",", ".")) * 1000)
                    cutting_data.append({
                        "partcode": partno,  # Parts поиск по артиклю
                        "count": count,
                        "length": l
                    })
                    if count > max_cnt and l > 0:
                        max_cnt = count
                        material_partno = partno

                #print(cutting_data)
                rolls.append(dict({"width": int(float(r.get("WIDTH").replace(",", ".")) * 1000),
                                   "height": int(float(r.get("HEIGHT").replace(",", ".")) * 1000),
                                   "material_partcode": material_partno,  # Parts поиск по артиклю
                                   "cutting_data": cutting_data  # Composition
                                   }))
            result = requests.post(f'http://127.0.0.1:8000/api/{sl[self.name_dow]}',
                                   json={'list': rolls})
            return
            #print(rolls)

        if "xlsx" in str(self.file):
            wb = load_workbook(f'media/{self.file}')


            sheet_name = wb.sheetnames[0]
            sheet = wb[sheet_name]

            parts = []
            for row in sheet.rows:
                if row[1].value == "Код" or row[1].value == None:
                    continue

                partcode = str(row[1].value).split("_")[0]
                l = 0
                cnt = 0
                if row[5].value != "-":
                    l = float(row[5].value.replace(",", "."))
                    cnt = int(float(row[6].value.replace(",", ".")) / l)
                else:
                    cnt = int(float(row[6].value.replace(",", ".")))

                l = int(l * 1000)

                parts.append({
                    "partcode": partcode,
                    "count": cnt,
                    "length": l
                })
            #print(parts)
            #return
            result = requests.post(f'http://127.0.0.1:8000/api/{sl[self.name_dow]}',
                                   json={'list': parts})
        else:

            with open(f"media/{self.file}", "r", encoding="utf8") as file1:
                lines = file1.read().splitlines()
            for line in lines:
                line = line.replace("(", "")
                line = line.replace("),", "")
                line = line.replace(")", "")
                line = line.replace(";", "")
                line = line.replace("'", "")
                line = line.replace("RAL ", "")
                #print(line.split(", "))
                sp.append(line.split(", "))
                #print(sp)
            #sp = json.dumps(sp)
                    #print(line.strip())
            result = requests.post(f'http://127.0.0.1:8000/api/{sl[self.name_dow]}',
                                   json={'list': sp})
        #print(result.text)






